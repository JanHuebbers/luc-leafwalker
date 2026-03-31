#!/usr/bin/env python3
"""
luc_empty_input_generator.py

Create template Excel workbook for luciferase leaf-assay input.

Sheets: 1..NE (or custom sheet names)
Each sheet columns:
  NLuc, CLuc, Replicate, Volume, Area

Rows:
  replicate blocks repeating the NLuc/CLuc pair list for each replicate.
  Optionally insert a blank spacer row between replicate blocks.

Notes:
- nluc_parts and cluc_parts must be same length and define row-wise pairings.
- normalizer must be present in nluc_parts or cluc_parts (validated), but is not suffixed.
"""

import argparse
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def parse_list(s: str) -> List[str]:
    items = [x.strip() for x in s.split(",") if x.strip()]
    if not items:
        raise ValueError("List is empty.")
    return items


def autosize_columns(ws, max_width: int = 45):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            best = max(best, len(str(v)))
        ws.column_dimensions[letter].width = min(max(best + 2, 10), max_width)


def make_workbook(
    out_path: str,
    num_experiments: int,
    num_replicates: int,
    nluc_parts: List[str],
    cluc_parts: List[str],
    normalizer: str,
    sheet_names: Optional[List[str]] = None,
    add_spacer_rows: bool = True,
):
    if num_experiments < 1:
        raise ValueError("num_experiments must be >= 1")
    if num_replicates < 1:
        raise ValueError("num_replicates must be >= 1")
    if len(nluc_parts) != len(cluc_parts):
        raise ValueError(
            f"nluc and cluc must have the same length, got {len(nluc_parts)} vs {len(cluc_parts)}"
        )
    if normalizer not in nluc_parts and normalizer not in cluc_parts:
        raise ValueError(
            f"Normalizer '{normalizer}' must be present in either nluc or cluc lists."
        )

    if sheet_names is None:
        sheet_names = [str(i) for i in range(1, num_experiments + 1)]
    if len(sheet_names) != num_experiments:
        raise ValueError("sheet_names length must equal num_experiments")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    header = ["NLuc", "CLuc", "Replicate", "Volume", "Area"]

    for name in sheet_names:
        ws = wb.create_sheet(title=name)

        # Header row
        ws.append(header)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        r = 2
        for rep in range(1, num_replicates + 1):
            for n, c in zip(nluc_parts, cluc_parts):
                ws.cell(row=r, column=1, value=n)
                ws.cell(row=r, column=2, value=c)
                ws.cell(row=r, column=3, value=rep)   # Replicate column
                # Volume/Area left blank
                r += 1

            if add_spacer_rows and rep != num_replicates:
                r += 1  # blank spacer row between blocks (left fully empty)

        ws.freeze_panes = "A2"
        autosize_columns(ws)

    wb.save(out_path)


def main():
    p = argparse.ArgumentParser(description="Create leaf-assay input XLSX template.")
    p.add_argument("--out", required=True, help="Output .xlsx path")
    p.add_argument("--num-experiments", type=int, required=True, help="Number of experiment sheets (NE)")
    p.add_argument("--num-replicates", type=int, required=True, help="Number of replicate blocks per sheet")

    p.add_argument(
        "--nluc",
        required=True,
        help='Comma-separated NLuc list, e.g. "AtMLO2,AtMLO2,AtMLO2"',
    )
    p.add_argument(
        "--cluc",
        required=True,
        help='Comma-separated CLuc list, e.g. "AtEXO70A1,AtEXO70A2,AtCAM2"',
    )
    p.add_argument(
        "--normalizer",
        required=True,
        help="Normalizer name (must appear in either --nluc or --cluc).",
    )
    p.add_argument(
        "--sheet-names",
        default=None,
        help='Optional comma-separated sheet names. If omitted: "1..NE".',
    )
    p.add_argument(
        "--no-spacers",
        action="store_true",
        help="Do not insert blank spacer rows between replicate blocks.",
    )

    args = p.parse_args()
    nluc_parts = parse_list(args.nluc)
    cluc_parts = parse_list(args.cluc)
    sheet_names = parse_list(args.sheet_names) if args.sheet_names else None

    make_workbook(
        out_path=args.out,
        num_experiments=args.num_experiments,
        num_replicates=args.num_replicates,
        nluc_parts=nluc_parts,
        cluc_parts=cluc_parts,
        normalizer=args.normalizer,
        sheet_names=sheet_names,
        add_spacer_rows=(not args.no_spacers),
    )


if __name__ == "__main__":
    main()
